import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, Alignment
import openpyxl
from openpyxl import load_workbook
from tkinter import messagebox
import datetime

# 创建Tkinter窗口
root = tk.Tk()
root.title("超长工单处理工具")
root.geometry("400x200")

# 创建一个Frame作为整个界面的容器，并使用pack布局来居中
frame = tk.Frame(root)
frame.pack(expand=True, fill="both")

# 创建一个Frame用于放置待更新超长工单的Label和Entry
updata_excel_frame = tk.Frame(frame)
updata_excel_frame.pack(side="top", fill="x", padx=10, pady=10, anchor="w")

def updata_excel_selectPath():
    path1.set(filedialog.askopenfilename())

path1 = tk.StringVar()
tk.Label(updata_excel_frame, text="待更新超长工单:").pack(side="left", padx=10, pady=10, anchor="w", fill="x", expand=True)
tk.Entry(updata_excel_frame, textvariable=path1).pack(side="left", fill="x", expand=True, padx=10, pady=10, anchor="w")
tk.Button(updata_excel_frame, text="文件路径选择", command=updata_excel_selectPath).pack(side="left", padx=10, pady=10, anchor="w", fill="x", expand=True)


# 创建一个Frame用于放置原超长工单的Label和Entry
Original_excel_frame = tk.Frame(frame)
Original_excel_frame.pack(side="top", fill="x", padx=10, pady=10, anchor="w")

def base_excel_selectPath():
    path2.set(filedialog.askopenfilename())

path2 = tk.StringVar()
tk.Label(Original_excel_frame, text="原超长工单:").pack(side="left", padx=10, pady=10, anchor="w", fill="x", expand=True) 
tk.Entry(Original_excel_frame, textvariable=path2).pack(side="left", fill="x", expand=True, padx=10, pady=10, anchor="w")
tk.Button(Original_excel_frame, text="文件路径选择", command=base_excel_selectPath).pack(side="left", padx=10, pady=10, anchor="w", fill="x", expand=True)


def run():
    try:
        workbook1 = load_workbook(path1.get())
        workbook2 = load_workbook(path2.get())

        all_sheet_names = workbook1.sheetnames
        # 创建一个新的 sheet 表格
        All_updates_sheet = workbook1.create_sheet("All updates")

        # 添加表头行
        titles = [cell.value for cell in workbook1[all_sheet_names[0]][1][:14]]
        titles[13] = "Time_diff"  # 将第13列的列名设置为 "time_diff"
        All_updates_sheet.append(titles)

        # 循环遍历所有 sheet 页
        for sheet_name in all_sheet_names:
            sheet = workbook1[sheet_name]
            # 将表格的数据保存在 rows 中
            rows = [[cell.value for cell in row][:14] for row in sheet.iter_rows(min_row=2)]
            # 将数据复制到新的 sheet 页中
            for row in rows:
                All_updates_sheet.append(row)

        # 删除所有 sheet 页，只保留新 sheet 页
        for sheet_name in all_sheet_names:
            if sheet_name != "All updates":
                workbook1.remove(workbook1[sheet_name])

        #新建一个名为待写入的sheet页
        new_sheet1 = workbook1.create_sheet("待写入")

        # 获取 "All updates" 工作表的表头
        header = [cell.value for cell in All_updates_sheet[1]]

        # 将表头添加到 "待写入" 工作表
        new_sheet1.append(header)

        # 将 All updates里`Task Created Date` 列的值转换为 datetime 格式，格式为"%Y/%m/%d %H:%M:%S"，并计算与当前时间的差值，写入time_diff列
        for index, row in enumerate(All_updates_sheet.iter_rows(min_row=2)):
            if row[12].value: 
               row[12].value = datetime.datetime.strptime(row[12].value, "%d/%m/%Y %H:%M:%S")
               row[13].value = datetime.datetime.now() - row[12].value
        

        for index, row in enumerate(All_updates_sheet.iter_rows(min_row=2)):
            if row[7].value != 'CCTV':
                if row[9].value == 'Processing' or row[9].value == 'Suspend':
                    if 'MIN'in row[11].value or 'NCR' in row[11].value or 'SLZ' in row[11].value or 'VIS' in row[11].value or 'NLZ' in row[11].value:
                        time_tiff = datetime.datetime.now() - row[12].value
                        if time_tiff.days > 5:
                            row[13].value = time_tiff.days
                             # 将每行的值赋给新的行
                            new_row = [cell.value for cell in row]
                            new_sheet1.append(new_row)


        # 保存工作簿
        workbook1.save(path1.get())

        #创建一个字典用于存储原超长工单的中的Ticket No.列的值
        Original_data_ticket = {}
        Original_sheet = workbook2["超长工单总表"]
        for index , row in Original_sheet.iter_rows(min_row=2):
            Original_data_ticket[index] = {
                "Original_Ticket": row[0].value
            }


        #创建一个字典用于存储待写入的中的Ticket No.列的值
        TB_write_all_data = {}
        TB_write_sheet = workbook1["待写入"]
        # 获取列名，假设它们在第一行
        column_names = [cell.value for cell in TB_write_sheet[1]]
        for index, row in enumerate(TB_write_sheet.iter_rows(min_row=2), start=2):
            TB_write_all_data[index] = {column_names[i]: cell.value for i, cell in enumerate(row)}

        #现在对比两个字典，如果原超长工单中的Ticket No.列的值与待写入的中的Ticket No.列的值存在重复，则将待写入字典中的该行数据删除，其余数据按照列名写入到原超长工单中
        for index1, (row1,) in enumerate(Original_sheet.iter_rows(min_row=2), start=2):
            for index2, row2 in enumerate(TB_write_sheet.iter_rows(min_row=2), start=2):
                if row1[0].value == row2[0].value:
                    TB_write_all_data.pop(index2, None)

        #将待写入字典中的数据按照列名写入到原超长工单中
        #打印出待写入字典中的数据
        print(TB_write_all_data)

        for index, row in TB_write_sheet.iter_rows(min_row=2):

            # 将每行的值赋给新的行
            new_row = [TB_write_all_data[index][column_name] for column_name in column_names]
            Original_sheet.append(new_row)

        # 保存工作簿
        workbook2.save(path2.get())


        # 弹出提示框，提示程序运行成功
        tk.messagebox.showinfo("提示", "程序运行成功！")

    except Exception as e:
        # 弹出错误提示框，提示程序运行出错
        tk.messagebox.showerror("错误", f"程序运行出错：{str(e)}")
        print(e)

    
# 创建运行按钮并使用pack布局居中
run_button = tk.Button(root, text="开始更新超长工单", command=run)
run_button.pack(side="top", pady=10)

# 运行Tkinter主循环
root.mainloop()